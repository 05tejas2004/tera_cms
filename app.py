from flask import Flask, render_template, redirect, url_for, request, flash
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask import Flask, render_template, request, redirect, url_for, flash, Response
from datetime import datetime
from fpdf import FPDF
import os
from sqlalchemy import func
import pandas as pd
from werkzeug.utils import secure_filename
from flask_login import current_user
from flask_login import login_user
from datetime import timedelta
import uuid
import json
from flask import jsonify


from flask import request, make_response, redirect, url_for
from datetime import datetime
from flask import Flask, render_template, request, make_response, redirect, url_for
from flask_socketio import SocketIO, emit # <-- Add these imports
from firebase_admin import messaging

import json
from flask import Flask, request, jsonify, render_template
from pywebpush import webpush, WebPushException

import io
from flask import render_template_string, make_response, request, redirect, url_for
from flask_login import login_required, current_user
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa
from datetime import datetime

app = Flask(__name__)
app.config['SECRET_KEY'] = 'tera-secret-key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://tera_cms_user:nvNoXUd6gShjT1emhtVdU1v90iRmFRJw@dpg-d8laqtv7f7vs73fqv1ig-a.oregon-postgres.render.com/tera_cms'
app.config['UPLOAD_FOLDER'] = 'static/uploads'

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'


# Replace your old socketio line with this one:
socketio = SocketIO(
    app, 
    cors_allowed_origins="*", 
    logger=True,          # Forces Flask to print WebSocket events to your terminal
    engineio_logger=True, # Prints low-level packet transfers
    ping_timeout=60, 
    ping_interval=25
)

VAPID_PRIVATE_KEY = "YOUR_VAPID_PRIVATE_KEY_HERE"
VAPID_CLAIMS = {
    "sub": "mailto:admin@yourdomain.com"
}

def send_push_notification(token, title, body):
    message = messaging.Message(
        notification=messaging.Notification(
            title=title,
            body=body,
        ),
        token=token,
    )
    response = messaging.send(message)

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_filtered_complaints_dataset():
    status_filter = request.args.get('status', '').strip()
    search_filter = request.args.get('search_filter', '').strip().lower()
    
    valid_statuses = ['Pending', 'Active', 'Closed']
    if status_filter in valid_statuses:
        query = Complaint.query.filter(Complaint.status == status_filter)
    else:
        query = Complaint.query.filter(Complaint.status.in_(valid_statuses))
        
    complaints = query.order_by(Complaint.created_at.desc()).all()
    all_users = User.query.all()
    user_dict = {u.id: u for u in all_users}
    
    if search_filter:
        filtered = []
        for c in complaints:
            short_id = (str(c.complaint_id)[-7:] if c.complaint_id else '').lower()
            title = (str(c.title) if c.title else '').lower()
            category = (str(c.category) if c.category else '').lower()
            location = (str(c.location) if c.location else '').lower()
            status = (str(c.status) if c.status else '').lower()
            
            if (search_filter in short_id or search_filter in title or 
                search_filter in category or search_filter in location or search_filter in status):
                filtered.append(c)
        complaints = filtered
        
    return complaints, user_dict


# --- Models ---
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100))
    username = db.Column(db.String(100), unique=True)
    password = db.Column(db.String(100))
    role = db.Column(db.String(20))
    is_approved = db.Column(db.Boolean, default=False)
    device_token = db.Column(db.String(100), nullable=True)  # ✅ Add this

class Complaint(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    complaint_id = db.Column(db.String(20), unique=True)
    title = db.Column(db.String(200))
    description = db.Column(db.Text)
    category = db.Column(db.String(50))
    priority = db.Column(db.String(20))
    location = db.Column(db.String(200))
    status = db.Column(db.String(20), default='Pending')
    created_at = db.Column(db.DateTime, default=datetime.now)
    assigned_to = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    accepted_at = db.Column(db.DateTime, nullable=True)  # Add THIS line
    assigned_at = db.Column(db.DateTime, nullable=True)
    resolved_at = db.Column(db.DateTime, nullable=True)
    resolved_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    resolution_remarks = db.Column(db.Text, nullable=True)
    worker_location = db.Column(db.String(100), nullable=True)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))

class ComplaintImage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    complaint_id = db.Column(db.Integer, db.ForeignKey('complaint.id'))
    filename = db.Column(db.String(200))

class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    complaint_id = db.Column(db.Integer, db.ForeignKey('complaint.id'))
    title = db.Column(db.String(200))
    message = db.Column(db.Text)
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    
class ComplaintLog(db.Model):
    __tablename__ = 'complaint_logs'
    id = db.Column(db.Integer, primary_key=True)
    complaint_id = db.Column(db.Integer, db.ForeignKey('complaint.id'), nullable=False)
    status = db.Column(db.String(50), nullable=False)
    changed_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    reason = db.Column(db.Text, nullable=True)
    changed_at = db.Column(db.DateTime, default=datetime.utcnow)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# --- Helper Functions ---
def get_resolution_time(created_at, resolved_at):
    if created_at and resolved_at:
        diff = resolved_at - created_at
        hours = diff.total_seconds() / 3600
        if hours < 1:
            return f"{int(diff.total_seconds() / 60)} min"
        elif hours < 24:
            return f"{int(hours)} hrs"
        return f"{int(hours / 24)} days"
    return "-"

# --- Routes ---
# Add this somewhere in your app (like after db.create_all())
@app.route('/api/save-subscription', methods=['POST'])
def save_subscription():
    subscription_payload = request.get_json()
    
    if not subscription_payload:
        return jsonify({"error": "Malformed subscription validation context"}), 400
        
    # Assume global session identifier context tracks current operator instances
    current_user_id = "user_123" 
    USER_SUBSCRIPTION_DB[current_user_id] = subscription_payload
    
    return jsonify({"status": "success", "message": "Device token written successfully."}), 200

def dispatch_push_alert(user_id, alert_title, alert_body, target_url):
    """ Internal microservice routine used to push standard payloads over WebPush protocol """
    user_subscription_info = USER_SUBSCRIPTION_DB.get(user_id)
    if not user_subscription_info:
        return False
        
    payload_data = json.dumps({
        "title": alert_title,
        "body": alert_body,
        "url": target_url,
        "icon": "/static/images/notification-icon.png"
    })
    
    try:
        webpush(
            subscription_info=user_subscription_info,
            data=payload_data,
            vapid_private_key=VAPID_PRIVATE_KEY,
            vapid_claims=VAPID_CLAIMS.copy()
        )
        return True
    except WebPushException as ex:
        print(f"Failed to dispatch remote push frames: {ex}")
        # Clean up expired subscription tokens if the browser client has uninstalled or rejected permissions
        if ex.response and ex.response.status_code in [404, 410]:
            del USER_SUBSCRIPTION_DB[user_id]
        return False
    
@app.route('/')
def home():
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        new_user = User(
            name=request.form['name'],
            username=request.form['username'],
            password=request.form['password'],
            role=request.form['role']
        )
        db.session.add(new_user)
        db.session.commit()
        flash('Registration submitted! Waiting for Admin approval.')
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Handle AJAX login (from JavaScript)
    if request.is_json:
        data = request.get_json()
        user = User.query.filter_by(username=data.get('username')).first()
        if user and user.password == data.get('password'):
            if not user.is_approved:
                return jsonify({'success': False, 'message': 'Account pending approval'}), 401
            
            remember_me = data.get('remember_me', False)
            login_user(user, remember=remember_me)
            
            if remember_me:
                if not user.device_token:
                    user.device_token = str(uuid.uuid4())
                    db.session.commit()
            
            return jsonify({'success': True, 'role': user.role})
        return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
    
    # Handle FORM submission (HTML form)
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form['username']).first()
        if user and user.password == request.form['password']:
            if not user.is_approved:
                flash('Account pending approval by Admin.')
                return redirect(url_for('login'))
            
            remember_me = request.form.get('remember_me', False)
            login_user(user, remember=remember_me)
            
            # ✅ If remember me, generate device token
            if remember_me:
                if not user.device_token:
                    user.device_token = str(uuid.uuid4())
                    db.session.commit()
            
            # ✅ FIX: Redirect to dashboard based on role (NOT JSON!)
            if user.role == 'Admin':
                return redirect(url_for('admin_dashboard'))
            elif user.role == 'Operator':
                return redirect(url_for('operator_dashboard'))
            elif user.role == 'HelpDesk':
                return redirect(url_for('helpdesk_dashboard'))
            elif user.role == 'FieldEngineer':
                return redirect(url_for('worker_dashboard'))
            else:
                return redirect(url_for('dashboard'))
        
        flash('Invalid credentials')
        return redirect(url_for('login'))
    
    return render_template('login.html')
@app.route('/auto_login', methods=['POST'])
def auto_login():
    """Auto login using device token"""
    data = request.get_json()
    device_token = data.get('device_token')
    
    if device_token:
        user = User.query.filter_by(device_token=device_token, is_approved=True).first()
        if user:
            login_user(user, remember=True)
            return jsonify({'success': True, 'role': user.role})
    
    return jsonify({'success': False}), 401


@app.route('/dashboard')
@login_required
def dashboard():

    if current_user.role == 'Admin':
        return redirect(url_for('admin_dashboard'))

    elif current_user.role == 'Operator':
        return redirect(url_for('operator_dashboard'))

    elif current_user.role == 'HelpDesk':
        return redirect(url_for('helpdesk_dashboard'))

    elif current_user.role == 'FieldEngineer':
        return redirect(url_for('worker_dashboard'))

    return f"Role not defined: {current_user.role}"

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# ============================================
# ADMIN DASHBOARD (UPDATED)
# ============================================

@app.route('/admin')
@login_required
def admin_dashboard():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    
    # Get all users
    all_users = User.query.all()
    
    # Get complaints
    complaints = Complaint.query.filter(
        Complaint.status.in_(['Pending', 'Active', 'Closed'])
    ).order_by(Complaint.created_at.desc()).all()
    
    # Stats
    total = len(complaints)
    pending = len([c for c in complaints if c.status == 'Pending'])
    active = len([c for c in complaints if c.status == 'Active'])
    closed = len([c for c in complaints if c.status == 'Closed'])
    
    # --- MOVED OPERATOR STATS HERE ---
    operators = User.query.filter_by(role='Operator', is_approved=True).count()
    # ----------------------------------
    
    # Pending users
    pending_users = User.query.filter_by(is_approved=False).all()
    
    # Filter by status
    filter_status = request.args.get('status')
    if filter_status:
        filtered_complaints = [c for c in complaints if c.status == filter_status]
    else:
        filtered_complaints = complaints
    
    return render_template('admin_dashboard.html',
                          total=total,
                          pending=pending,
                          active=active,
                          closed=closed,
                          all_users=all_users,
                          pending_users=pending_users,
                          complaints=complaints,
                          filtered_complaints=filtered_complaints,
                          operators=operators, # <-- Passed to the Admin Dashboard template
                          current_time=datetime.now())



# ============================================
# ADMIN - USER MANAGEMENT
# ============================================

@app.route('/admin/users')
@login_required
def user_management():
    if current_user.role != 'Admin': return redirect(url_for('dashboard'))
    
    all_users = User.query.all()
    pending_users = User.query.filter_by(is_approved=False).all()
    approved_users = User.query.filter_by(is_approved=True).all()
    
    # Count users by role
    admin_count = User.query.filter_by(role='Admin', is_approved=True).count()
    operator_count = User.query.filter_by(role='Operator', is_approved=True).count()
    helpdesk_count = User.query.filter_by(role='HelpDesk', is_approved=True).count()
    engineer_count = User.query.filter_by(role='FieldEngineer', is_approved=True).count()
    
    return render_template('user_management.html', 
                         all_users=all_users,
                         pending_users=pending_users,
                         approved_users=approved_users,
                         admin_count=admin_count,
                         operator_count=operator_count,
                         helpdesk_count=helpdesk_count,
                         engineer_count=engineer_count)

@app.route('/admin/delete_user/<int:user_id>')
@login_required
def admin_delete_user(user_id):
    if current_user.role != 'Admin': return redirect(url_for('dashboard'))
    
    user = User.query.get(user_id)
    if user:
        # Don't delete admin
        if user.role == 'Admin':
            flash('Cannot delete Admin user!', 'error')
        else:
            db.session.delete(user)
            db.session.commit()
            flash(f'User {user.username} deleted successfully!', 'success')
    
    return redirect(url_for('user_management'))

@app.route('/admin/toggle_user_status/<int:user_id>')
@login_required
def toggle_user_status(user_id):
    if current_user.role != 'Admin': return redirect(url_for('dashboard'))
    
    user = User.query.get(user_id)
    if user and user.role != 'Admin':
        user.is_approved = not user.is_approved
        db.session.commit()
        status = "approved" if user.is_approved else "deactivated"
        flash(f'User {user.username} {status}!', 'success')
    
    return redirect(url_for('user_management'))

@app.route('/admin/approve_user/<int:user_id>')
@login_required
def approve_user(user_id):
    if current_user.role != 'Admin': return redirect(url_for('dashboard'))
    user = User.query.get(user_id)
    user.is_approved = True
    db.session.commit()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/reject_user/<int:user_id>')
@login_required
def reject_user(user_id):
    if current_user.role != 'Admin': return redirect(url_for('dashboard'))
    user = User.query.get(user_id)
    db.session.delete(user)
    db.session.commit()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete_user/<int:user_id>')
@login_required
def delete_user(user_id):
    if current_user.role != 'Admin': return redirect(url_for('dashboard'))
    user = User.query.get(user_id)
    db.session.delete(user)
    db.session.commit()
    return redirect(url_for('admin_dashboard'))

# ============================================
# GENERATE PDF REPORT
# ============================================

@app.route('/generate_report')
@login_required
def generate_report():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
        
    # Detect the intended format type ('pdf' or 'excel')
    report_format = request.args.get('format', 'excel').lower()
    
    # Run your exact filter engine
    complaints, user_dict = get_filtered_complaints_dataset()
    current_time = datetime.now()
    
    # Standardized Global Headings list containing all your specified targets
    headers = [
        "Complaint ID", "Title", "Description", "Category", "Location", 
        "Created By", "Accepted By", "Resolved By", "Created", "Accepted", 
        "Time", "Reason"
    ]
    
    # --- BRANCH 1: EXCEL FORMAT ---
    if report_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Complaints Export"
        ws.views.sheetView[0].showGridLines = True
        
        SLATE_HEADER = "343A40"
        ZEBRA_LIGHT = "F8F9FA"
        
        font_title = Font(name="Calibri", size=16, bold=True, color="212529")
        font_sub = Font(name="Calibri", size=10, italic=True, color="6C757D")
        font_h = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_d = Font(name="Calibri", size=11, color="212529")
        
        fill_h = PatternFill(start_color=SLATE_HEADER, end_color=SLATE_HEADER, fill_type="solid")
        fill_z = PatternFill(start_color=ZEBRA_LIGHT, end_color=ZEBRA_LIGHT, fill_type="solid")
        border_thin = Border(
            left=Side(style="thin", color="DEE2E6"), right=Side(style="thin", color="DEE2E6"),
            top=Side(style="thin", color="DEE2E6"), bottom=Side(style="thin", color="DEE2E6")
        )
        
        ws["A1"] = "Complaint Management System - Dashboard Export"
        ws["A1"].font = font_title
        ws["A2"] = f"Generated: {current_time.strftime('%Y-%m-%d %H:%M')} | Data Snapshot"
        ws["A2"].font = font_sub
        
        # Draw Headers Dynamically
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.font = font_h
            cell.fill = fill_h
            cell.border = border_thin
            cell.alignment = Alignment(horizontal="center" if col_idx in [1, 9, 10, 11] else "left", vertical="center")
            
        for idx, c in enumerate(complaints, start=5):
            # Safe mapping lookup tracking attributes
            created_by_user = user_dict.get(c.created_by).name if (c.created_by and c.created_by in user_dict) else "-"
            accepted_by_user = user_dict.get(c.assigned_to).name if (c.assigned_to and c.assigned_to in user_dict) else "-"
            resolved_by_user = user_dict.get(c.resolved_by).name if (c.resolved_by and c.resolved_by in user_dict) else "-"
            
            created = c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else '-'
            accepted = c.assigned_at.strftime('%Y-%m-%d %H:%M') if c.assigned_at else '-'
            
            time_diff = "-"
            if c.created_at:
                ref_time = c.assigned_at if c.assigned_at else current_time
                diff_sec = (ref_time - c.created_at).total_seconds()
                hours = int(diff_sec // 3600)
                time_diff = f"{int(diff_sec // 60)}m" if hours < 1 else (f"{hours}h" if hours < 24 else f"{int(hours // 24)}d")
                
            desc = c.description or "-"
            reason = c.resolution_remarks or "-"
            
            # Map values precisely to standard header indices
            row_data = [
                str(c.complaint_id), str(c.title), desc, str(c.category), str(c.location),
                created_by_user, accepted_by_user, resolved_by_user, created, accepted, 
                time_diff, reason
            ]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=idx, column=col_idx, value=val)
                cell.font = font_d
                cell.border = border_thin
                if idx % 2 == 0:
                    cell.fill = fill_z
                cell.alignment = Alignment(horizontal="center" if col_idx in [1, 9, 10, 11] else "left", vertical="center")
                
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 11), 35)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        response = make_response(stream.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=complaints_dataset_export.xlsx'
        return response

    # --- BRANCH 2: PDF FORMAT ---
    elif report_format == 'pdf':
        html_template = """
        <!DOCTYPE html>
        <html>
        <head>
        <style>
            @page { size: A4 landscape; margin: 8mm; background-color: #ffffff; }
            body { font-family: Arial, sans-serif; color: #212529; font-size: 7.5pt; margin: 0; }
            .header { border-bottom: 2px solid #343a40; padding-bottom: 5px; margin-bottom: 12px; }
            h2 { margin: 0 0 4px 0; color: #343a40; font-size: 14pt; }
            .meta { color: #6c757d; font-size: 8pt; margin: 0; }
            table { width: 100%; border-collapse: collapse; margin-top: 5px; table-layout: fixed; }
            th { background-color: #343a40; color: white; font-weight: bold; border: 1px solid #dee2e6; padding: 5px; font-size: 7.5pt; text-align: left; }
            td { border: 1px solid #dee2e6; padding: 5px; font-size: 7pt; vertical-align: top; word-wrap: break-word; overflow: hidden; }
            tr:nth-child(even) td { background-color: #f8f9fa; }
            .center { text-align: center; }
            .badge { display: inline-block; padding: 1px 4px; font-size: 7pt; border-radius: 2px; color: white; background-color: #6c757d; }
            .bg-success { background-color: #198754; }
            .bg-info { background-color: #0dcaf0; color: #000; }
            .bg-warning { background-color: #ffc107; color: #000; }
        </style>
        </head>
        <body>
            <div class="header">
                <h2>Complaint Management System - Report Archive</h2>
                <p class="meta">Generated: {{ current_time.strftime('%Y-%m-%d %H:%M') }}</p>
            </div>
            <table>
                <thead>
                    <tr>
                        <th style="width: 8%;">Complaint ID</th>
                        <th style="width: 10%;">Title</th>
                        <th style="width: 14%;">Description</th>
                        <th style="width: 8%;">Category</th>
                        <th style="width: 8%;">Location</th>
                        <th style="width: 8%;">Created By</th>
                        <th style="width: 8%;">Accepted By</th>
                        <th style="width: 8%;">Resolved By</th>
                        <th style="width: 8%;">Created</th>
                        <th style="width: 8%;">Accepted</th>
                        <th style="width: 4%;">Time</th>
                        <th style="width: 8%;">Reason</th>
                    </tr>
                </thead>
                <tbody>
                    {% for c in complaints %}
                    <tr>
                        <td class="center"><strong>{{ c.complaint_id }}</strong></td>
                        <td>{{ c.title }}</td>
                        <td>{{ c.description or '-' }}</td>
                        <td><span class="badge">{{ c.category }}</span></td>
                        <td>{{ c.location }}</td>
                        
                        <td>
                            {% if c.created_by and c.created_by in user_dict %}
                                {{ user_dict[c.created_by].name }}
                            {% else %}
                                <span class="text-muted">-</span>
                            {% endif %}
                        </td>
                        
                        <td>
                            {% if c.assigned_to and c.assigned_to in user_dict %}
                                <span class="badge bg-success">{{ user_dict[c.assigned_to].name }}</span>
                            {% else %}
                                <span class="text-muted">-</span>
                            {% endif %}
                        </td>
                        
                        <td>
                            {% if c.resolved_by and c.resolved_by in user_dict %}
                                <span class="badge bg-info">{{ user_dict[c.resolved_by].name }}</span>
                            {% else %}
                                <span class="text-muted">-</span>
                            {% endif %}
                        </td>
                        
                        <td class="center">{{ c.created_at.strftime('%Y-%m-%d') if c.created_at else '-' }}</td>
                        <td class="center">{{ c.assigned_at.strftime('%Y-%m-%d') if c.assigned_at else '-' }}</td>
                        
                        <td class="center">
                            {% if c.created_at %}
                                {% set r_time = c.assigned_at if c.assigned_at else current_time %}
                                {% set diff = r_time - c.created_at %}
                                {% set hours = (diff.total_seconds() / 3600)|int %}
                                {% if hours < 1 %}
                                    <span class="badge bg-warning">{{ (diff.total_seconds() / 60)|int }}m</span>
                                {% elif hours < 24 %}
                                    <span class="badge bg-success">{{ hours }}h</span>
                                {% else %}
                                    <span class="badge bg-success">{{ (hours / 24)|int }}d</span>
                                {% endif %}
                            {% else %}
                                -
                            {% endif %}
                        </td>
                        
                        <td>
                            {% if c.resolution_remarks %}
                                {{ c.resolution_remarks.split(':')[-1].strip() if ':' in c.resolution_remarks else c.resolution_remarks }}
                            {% else %}
                                -
                            {% endif %}
                        </td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </body>
        </html>
        """
        rendered_html = render_template_string(html_template, complaints=complaints, user_dict=user_dict, current_time=current_time)
        pdf_stream = io.BytesIO()
        
        pisa_status = pisa.CreatePDF(io.StringIO(rendered_html), dest=pdf_stream)
        
        if pisa_status.err:
            return "Error generating PDF", 500
            
        pdf_stream.seek(0)
        response = make_response(pdf_stream.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=complaints_matrix_report.pdf'
        return response

# ============================================
# THEN, ADD THE ROUTE AFTER ALL OTHER ROUTES
# ============================================

@app.route('/admin/upload_users', methods=['GET', 'POST'])
@login_required
def upload_users():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected!', 'error')
            return redirect(request.url)
        
        file = request.files['file']
        
        if file.filename == '':
            flash('No file selected!', 'error')
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            file.save(filepath)
            
            try:
                # Read Excel file
                df = pd.read_excel(filepath)
                
                # Required columns
                required_cols = ['name', 'username', 'password', 'role']
                
                # Check if columns exist
                missing_cols = [col for col in required_cols if col not in df.columns]
                if missing_cols:
                    flash(f'Missing columns: {missing_cols}', 'error')
                    return redirect(request.url)
                
                # Create/Update users
                created_count = 0
                updated_count = 0
                skipped_count = 0
                
                for index, row in df.iterrows():
                    name = str(row['name']).strip()
                    username = str(row['username']).strip()
                    password = str(row['password']).strip()
                    role = str(row['role']).strip()
                    
                    # Validate role
                    if role not in ['Admin', 'Operator', 'HelpDesk', 'FieldEngineer']:
                        skipped_count += 1
                        continue
                    
                    # Check if user exists
                    existing_user = User.query.filter_by(username=username).first()
                    
                    if existing_user:
                        # Update existing user (only name and role)
                        existing_user.name = name
                        existing_user.role = role
                        updated_count += 1
                    else:
                        # Create new user - use password directly
                        new_user = User(
                            name=name,
                            username=username,
                            password=password,  # Direct assignment
                            role=role,
                            is_approved=True
                        )
                        db.session.add(new_user)
                        created_count += 1
                
                db.session.commit()
                os.remove(filepath)
                
                flash(f'Successfully! Created: {created_count}, Updated: {updated_count}, Skipped: {skipped_count}', 'success')
                return redirect(url_for('admin_dashboard'))
            
            except Exception as e:
                flash(f'Error reading file: {str(e)}', 'error')
                return redirect(request.url)
    
    return render_template('upload_users.html')
# ============================================
# ADMIN - STATISTICS
# ============================================

@app.route('/admin/statistics')
@login_required
def admin_statistics():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    
    # === User Statistics ===
    total_users = User.query.count()
    approved_users = User.query.filter_by(is_approved=True).count()
    pending_users = User.query.filter_by(is_approved=False).count()
    engineers = User.query.filter_by(role='FieldEngineer', is_approved=True).count()
    # REMOVED: operators query stripped from here
    helpdesk = User.query.filter_by(role='HelpDesk', is_approved=True).count()
    admins = User.query.filter_by(role='Admin', is_approved=True).count()
    
    # === Complaint Statistics - ONLY Pending, Active, Closed ===
    total_complaints = Complaint.query.filter(
        Complaint.status.in_(['Pending', 'Active', 'Closed'])
    ).count()
    
    pending = Complaint.query.filter_by(status='Pending').count()
    active = Complaint.query.filter_by(status='Active').count()
    closed = Complaint.query.filter_by(status='Closed').count()
    resolved = 0  # Not showing Resolved
    rejected = 0  # Not showing Rejected
    
    # === Category Statistics - ONLY Pending, Active, Closed ===
    category_counts = {}
    valid_complaints = Complaint.query.filter(
        Complaint.status.in_(['Pending', 'Active', 'Closed'])
    ).all()
    
    for c in valid_complaints:
        if c.category:
            cat = c.category
            if cat in category_counts:
                category_counts[cat] += 1
            else:
                category_counts[cat] = 1
    
    # === Priority Statistics - ONLY Pending, Active, Closed ===
    high_priority = Complaint.query.filter(
        Complaint.status.in_(['Pending', 'Active', 'Closed']),
        Complaint.priority == 'High'
    ).count()
    
    medium_priority = Complaint.query.filter(
        Complaint.status.in_(['Pending', 'Active', 'Closed']),
        Complaint.priority == 'Medium'
    ).count()
    
    low_priority = Complaint.query.filter(
        Complaint.status.in_(['Pending', 'Active', 'Closed']),
        Complaint.priority == 'Low'
    ).count()
    
    # === Engineer Performance - ONLY Pending, Active, Closed ===
    engineer_list = []
    engineers_users = User.query.filter_by(role='FieldEngineer', is_approved=True).all()
    
    for eng in engineers_users:
        # Count only Pending, Active, Closed
        assigned_count = Complaint.query.filter(
            Complaint.assigned_to == eng.id,
            Complaint.status.in_(['Pending', 'Active', 'Closed'])
        ).count()
        
        closed_count = Complaint.query.filter_by(
            assigned_to=eng.id, 
            status='Closed'
        ).count()
        
        active_count = Complaint.query.filter_by(
            assigned_to=eng.id, 
            status='Active'
        ).count()
        
        engineer_list.append({
            'name': eng.name,
            'assigned': assigned_count,
            'resolved': closed_count,
            'active': active_count
        })
    
    return render_template('admin_statistics.html',
                           total_users=total_users,
                           approved_users=approved_users,
                           pending_users=pending_users,
                           engineers=engineers,
                           # REMOVED: operators mapping omitted
                           helpdesk=helpdesk,
                           admins=admins,
                           total_complaints=total_complaints,
                           pending=pending,
                           active=active,
                           resolved=resolved,
                           closed=closed,
                           rejected=rejected,
                           category_counts=category_counts,
                           high_priority=high_priority,
                           medium_priority=medium_priority,
                           low_priority=low_priority,
                           engineer_performance=engineer_list)

# ============================================
# ADMIN - UPDATE COMPLAINT STATUS
# ============================================

@app.route('/admin/update_status/<int:comp_id>', methods=['POST'])
@login_required
def admin_update_status(comp_id):
    if current_user.role != 'Admin': 
        return redirect(url_for('dashboard'))
    
    complaint = Complaint.query.get(comp_id)
    if not complaint:
        flash('Complaint not found!', 'error')
        return redirect(url_for('admin_dashboard'))
    
    new_status = request.form.get('status', '').strip()
    
    if new_status:
        old_status = complaint.status
        complaint.status = new_status
        
        if new_status == 'Closed':
            complaint.resolved_at = datetime.now()
        
        db.session.commit()

        # --- REAL-TIME BROADCAST TRIGGER FOR ADMIN ---
        try:
            socketio.emit('global_notification', {
                'title': '売 Status Updated by Admin',
                'message': f"Complaint ID ...{str(complaint.complaint_id)[-7:]} has been changed from '{old_status}' to '{new_status}' by System Admin."
            }, broadcast=True)
        except Exception as e:
            print(f"WebSocket Admin broadcast error: {e}")
        # -----------------------------------------------

        flash(f'Complaint {complaint.complaint_id} status updated from {old_status} to {new_status}!', 'success')
    
    return redirect(url_for('admin_dashboard'))
# ============================================
# ADMIN - DOWNLOAD ALL USERS EXCEL
# ============================================


@app.route('/admin/download_users_excel')
@login_required
def download_users_excel():
    if current_user.role != 'Admin':
        return redirect(url_for('dashboard'))
    
    # Get all users
    users = User.query.all()
    
    # Create DataFrame
    data = []
    for user in users:
        data.append({
            'name': user.name,
            'username': user.username,
            'password': user.password,  # Note: This shows hashed password
            'role': user.role
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel file
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'all_users.xlsx')
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    
    df.to_excel(filepath, index=False)
    
    # Return file for download
    from flask import send_file
    
    return send_file(
        filepath,
        as_attachment=True,
        download_name='all_users.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ============================================
# ADMIN - VIEW COMPLAINT DETAILS
# ============================================

@app.route('/admin/view_complaint/<int:comp_id>')
@login_required
def admin_view_complaint(comp_id):
    if current_user.role != 'Admin': return redirect(url_for('dashboard'))
    
    complaint = Complaint.query.get(comp_id)
    all_users = User.query.all()
    images = ComplaintImage.query.filter_by(complaint_id=complaint.id).all()
    
    return render_template('admin_view_complaint.html', 
                         complaint=complaint,
                         all_users=all_users,
                         images=images)

@app.route('/complaint/view/<int:id>')
@login_required
def view_single_complaint(id):
    complaint = Complaint.query.get_or_404(id)
    creator = User.query.get(complaint.created_by) if complaint.created_by else None
    resolver = User.query.get(complaint.resolved_by) if complaint.resolved_by else None
    
    # CRITICAL: Fetch logs from oldest to newest to trace the history cleanly
    logs = ComplaintLog.query.filter_by(complaint_id=complaint.id).order_by(ComplaintLog.changed_at.asc()).all()
    
    # Create a quick directory map to resolve usernames dynamically inside the log table
    all_users = {u.id: u.name for u in User.query.all()}
    
    return render_template('complaint_view.html', 
                           complaint=complaint, 
                           creator=creator, 
                           resolver=resolver,
                           logs=logs,
                           all_users=all_users)
# ============================================
# OPERATOR DASHBOARD
# ============================================

@app.route('/operator')
@login_required
def operator_dashboard():
    # Role-based validation gate check
    if current_user.role != 'Operator': 
        return redirect(url_for('dashboard'))
    
    # Base Collections
    complaints = Complaint.query.order_by(Complaint.created_at.desc()).all()
    resolved_complaints = Complaint.query.filter_by(status='Resolved').all()
    all_users = User.query.all()
    
    # ----------------------------------------------------
    # NEW KPI METRICS QUERIES
    # ----------------------------------------------------
    total_count = Complaint.query.count()
    active_count = Complaint.query.filter_by(status='Active').count()
    resolved_count = Complaint.query.filter_by(status='Resolved').count()
    closed_count = Complaint.query.filter_by(status='Closed').count()
    # ----------------------------------------------------
    
    # Fetch image dictionary mappings safely
    complaint_images = {}
    for c in complaints:
        images = ComplaintImage.query.filter_by(complaint_id=c.id).all()
        complaint_images[c.id] = images
    
    return render_template('operator_dashboard.html', 
                           complaints=complaints,
                           resolved_complaints=resolved_complaints,
                           all_users=all_users,
                           complaint_images=complaint_images,
                           get_resolution_time=get_resolution_time,
                           # Pass metrics keys context down to view template layer
                           total_count=total_count,
                           active_count=active_count,
                           resolved_count=resolved_count,
                           closed_count=closed_count)

@app.route('/operator/stats')
@login_required
def operator_stats():
    # Role gate check
    if current_user.role != 'Operator':
        return redirect(url_for('dashboard'))

    # 1. Broad High-Level Statistics Counters
    total_complaints = Complaint.query.count()
    pending_count = Complaint.query.filter_by(status='Pending').count()
    active_count = Complaint.query.filter_by(status='Active').count()
    resolved_count = Complaint.query.filter_by(status='Resolved').count()
    closed_count = Complaint.query.filter_by(status='Closed').count()

    # 2. Category Breakdowns
    category_data = db.session.query(
        Complaint.category, func.count(Complaint.id)
    ).group_by(Complaint.category).all()
    
    categories = [row[0] for row in category_data if row[0]]
    category_counts = [row[1] for row in category_data if row[0]]

    # 3. Priority Breakdowns
    priority_data = db.session.query(
        Complaint.priority, func.count(Complaint.id)
    ).group_by(Complaint.priority).all()
    
    priorities = [row[0] for row in priority_data if row[0]]
    priority_counts = [row[1] for row in priority_data if row[0]]

    # 4. Location-based Frequency Breakdowns (Top 5 locations)
    location_data = db.session.query(
        Complaint.location, func.count(Complaint.id)
    ).group_by(Complaint.location).order_by(func.count(Complaint.id).desc()).limit(5).all()
    
    locations = [row[0] for row in location_data if row[0]]
    location_counts = [row[1] for row in location_data if row[0]]

    return render_template(
        'operator_stats.html',
        total=total_complaints,
        pending=pending_count,
        active=active_count,
        resolved=resolved_count,
        closed=closed_count,
        categories=categories,
        category_counts=category_counts,
        priorities=priorities,
        priority_counts=priority_counts,
        locations=locations,
        location_counts=location_counts
    )
# ============================================
# OPERATOR - VERIFY COMPLAINT (Separate Page)
# ============================================

@app.route('/operator/verify/<int:comp_id>')
@login_required
def verify_complaint(comp_id):
    if current_user.role != 'Operator': return redirect(url_for('dashboard'))
    
    complaint = Complaint.query.get(comp_id)
    if not complaint:
        flash('Complaint not found!', 'error')
        return redirect(url_for('operator_dashboard'))
    
    # Get assigned user details
    assigned_user = User.query.get(complaint.assigned_to) if complaint.assigned_to else None
    
    # Get resolution images
    images = ComplaintImage.query.filter_by(complaint_id=complaint.id).all()
    
    # Get all users for display
    all_users = User.query.all()
    
    return render_template('verify_complaint.html', 
                         complaint=complaint,
                         assigned_user=assigned_user,
                         images=images,
                         all_users=all_users,
                         get_resolution_time=get_resolution_time)

@app.route('/operator/update/<int:comp_id>', methods=['POST'])
@login_required
def tech_support_update_complaint(comp_id):
    # Ensure authorization
    if current_user.role not in ['TechnicalSupport', 'Operator']:
        return redirect(url_for('dashboard'))
        
    comp = Complaint.query.get_or_404(comp_id)
    old_status = comp.status
    new_status = request.form.get('status')
    user_reason = request.form.get('reason')
    
    # Track who updated this complaint
    comp.assigned_to = current_user.id  # Links the "Resolved By" lookup
    comp.status = new_status
    
    if new_status == 'Closed':
        comp.closed_at = datetime.now()
    elif new_status == 'Active':
        comp.assigned_at = datetime.now() 
    
    # Save the structured log string format
    timestamp = datetime.now().strftime('%d-%m-%Y %H:%M')
    comp.resolution_remarks = f"[{timestamp}] {current_user.name} changed status to {new_status}: {user_reason}"
    
    db.session.commit()

    # --- REAL-TIME BROADCAST TRIGGER FOR OPERATORS ---
    try:
        socketio.emit('global_notification', {
            'title': '売 Status Updated by Operator',
            'message': f"Complaint ID ...{str(comp.complaint_id)[-7:]} has been changed from '{old_status}' to '{new_status}' by {current_user.name}."
        }, broadcast=True)
    except Exception as e:
        print(f"WebSocket Operator broadcast error: {e}")
    # --------------------------------------------------
    
    flash(f'Complaint status updated to {new_status} successfully.', 'success')
    return redirect(url_for('operator_dashboard'))
# ============================================
# OPERATOR - PHOTO VIEWER (Full Screen)
# ============================================

@app.route('/operator/photos/<int:comp_id>')
@login_required
def photo_viewer(comp_id):
    if current_user.role not in ['Operator', 'HelpDesk', 'Admin']:
        return redirect(url_for('dashboard'))
    
    complaint = Complaint.query.get(comp_id)
    if not complaint:
        flash('Complaint not found!', 'error')
        return redirect(url_for('operator_dashboard'))
    
    images = ComplaintImage.query.filter_by(complaint_id=complaint.id).all()
    all_complaints = Complaint.query.filter(Complaint.id != complaint.id).order_by(Complaint.id.desc()).all()
    
    # Fix: Convert SQLAlchemy objects to list of filenames (strings)
    image_list = [img.filename for img in images]
    
    return render_template('photo_viewer.html', 
                         complaint=complaint,
                         image_list=image_list,  # Use this instead of images
                         all_complaints=all_complaints)
# ============================================
# HELP DESK DASHBOARD
# ============================================

@app.route('/helpdesk')
@login_required
def helpdesk_dashboard():
    if current_user.role != 'HelpDesk': 
        return redirect(url_for('dashboard'))
    
    # 1. Base Query sorted from latest to oldest (Newest on top)
    complaints = Complaint.query.order_by(Complaint.created_at.desc()).all()
    
    # Fetch engineers for any dropdown assignment logic
    engineers = User.query.filter_by(role='FieldEngineer', is_approved=True).all()
    
    # Fetch all users so the template can match names to IDs seamlessly
    all_users = User.query.all()
    
    # Get filter status parameter from URL string
    filter_status = request.args.get('status')
    
    # 2. Extract filtered list while preserving the database desc() sorting order
    if filter_status:
        filtered_complaints = [c for c in complaints if c.status == filter_status]
    else:
        filtered_complaints = complaints
    
    return render_template('helpdesk_dashboard.html', 
                           complaints=complaints, 
                           engineers=engineers,
                           filtered_complaints=filtered_complaints,
                           all_users=all_users)
# ============================================
# FIELD ENGINEER DASHBOARD
# ============================================

@app.route('/worker')
@login_required
def worker_dashboard():
    if current_user.role != 'FieldEngineer': 
        return redirect(url_for('dashboard'))
    
    # Get all engineers for display
    engineers = User.query.filter_by(role='FieldEngineer', is_approved=True).all()
    
    # Get filter status
    filter_status = request.args.get('status')
    
    # All pending (available)
    pending_complaints = Complaint.query.filter_by(status='Pending').all()
    
    # My complaints
    my_complaints = Complaint.query.filter_by(assigned_to=current_user.id).all()
    my_pending = [c for c in my_complaints if c.status == 'Pending']
    my_active = [c for c in my_complaints if c.status == 'Active']
    my_resolved = [c for c in my_complaints if c.status == 'Resolved']
    my_history = [c for c in my_complaints if c.status in ['Closed', 'Rejected']]
    
    return render_template('worker_dashboard.html',
                         pending_complaints=pending_complaints,
                         my_complaints=my_complaints,
                         my_pending=my_pending,
                         my_active=my_active,
                         my_resolved=my_resolved,
                         my_history=my_history,
                         engineers=engineers)


# ============================================
# NOTIFICATIONS
# ============================================

@app.route('/notifications')
@login_required
def notifications():
    notifs = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.created_at.desc()).all()
    return render_template('notifications.html', notifications=notifs)

@app.route('/mark_notif_read/<int:notif_id>')
@login_required
def mark_notif_read(notif_id):
    notif = Notification.query.get(notif_id)
    if notif and notif.user_id == current_user.id:
        notif.is_read = True
        db.session.commit()
    return redirect(url_for('notifications'))

# ============================================
# REGISTER COMPLAINT
# ============================================



@app.route('/register_complaint', methods=['GET', 'POST'])
@login_required
def register_complaint():
    if current_user.role not in ['HelpDesk', 'Operator']:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        categories = request.form.getlist('category[]')
        new_comp = Complaint(
            complaint_id=f"CMP-{datetime.now().strftime('%Y%m%d%H%M%S')}",
            title=request.form['title'],
            description=request.form['description'],
            category=", ".join(categories),
            priority=request.form['priority'],
            location=request.form['location'],
            created_by=current_user.id,
            status='Pending'
        )

        db.session.add(new_comp)
        db.session.flush() # Generates the new_comp.id dynamically
        
        # --- AUDIT TRAIL LOGGING: INITIAL COMPLAINT REGISTRATION ---
        initial_log = ComplaintLog(
            complaint_id=new_comp.id,
            status='Pending',
            changed_by=current_user.id,
            reason="Complaint successfully registered in the ecosystem."
        )
        db.session.add(initial_log)
        
        # Notify ALL field engineers via database layout records
        engineers = User.query.filter_by(role='FieldEngineer', is_approved=True).all()
        for engineer in engineers:
            notif = Notification(
                user_id=engineer.id,
                complaint_id=new_comp.id,
                title="New Complaint Available",
                message=f"New complaint: {new_comp.title} at {new_comp.location}. Category: {new_comp.category}"
            )
            db.session.add(notif)
        
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"Database commit exception: {e}")
            flash('An error occurred while saving the complaint.', 'error')
            return redirect(url_for('helpdesk_dashboard'))
        
        # --- GLOBAL SOCKET REAL-TIME BROADCAST WITH SOUND ---
        try:
            socketio.emit('global_notification', {
                'title': '🔔 New Complaint Registered!',
                'message': f"Complaint {new_comp.complaint_id[-7:]}: '{new_comp.title}' has been reported at {new_comp.location}."
            }, broadcast=True)
        except Exception as e:
            print(f"WebSocket broadcast exception: {e}")
        # ----------------------------------------------------
        
        if engineers:
            flash(f'Complaint registered! All {len(engineers)} engineers notified.')
        else:
            flash('Complaint registered! No engineers available.')
        
        if current_user.role == 'Operator':
            return redirect(url_for('operator_dashboard'))
        else:
            return redirect(url_for('helpdesk_dashboard'))
    
    return render_template('register_complaint.html')


@app.route('/update_complaint_status/<int:comp_id>', methods=['POST'])
@login_required
def update_complaint_status(comp_id):
    if current_user.role not in ['Admin', 'HelpDesk', 'Operator', 'FieldEngineer']:
        return redirect(url_for('dashboard'))
        
    complaint = Complaint.query.get_or_404(comp_id)
    old_status = complaint.status
    new_status = request.form.get('status')
    
    # Capture the context reasoning text field (Fallback check for 'remarks' or 'reason' forms)
    remarks = request.form.get('remarks', request.form.get('reason', '')).strip()
    
    if new_status and (new_status != old_status or remarks):
        complaint.status = new_status
        complaint.resolved_by = current_user.id  # Dynamically logs who performed the status modification
        
        # Format closing text notes logs context
        log_reason = remarks if remarks else f"Status manually adjusted to {new_status} by {current_user.role}."
        complaint.resolution_remarks = f"{current_user.role}: {log_reason}"
        
        if new_status in ['Active', 'Closed']:
            complaint.assigned_at = datetime.now()
            
        # --- AUDIT TRAIL LOGGING: TRACK SYSTEM STATUS CHANGE ---
        new_log = ComplaintLog(
            complaint_id=complaint.id,
            status=new_status,
            changed_by=current_user.id,
            reason=log_reason
        )
        db.session.add(new_log)
        
        try:
            db.session.commit()
            flash(f"Complaint status updated to {new_status} successfully!")
        except Exception as e:
            db.session.rollback()
            print(f"Database update exception: {e}")
            flash('An error occurred while updating the complaint status.', 'error')
        
        # --- GLOBAL SOCKET REAL-TIME BROADCAST FOR STATUS CHANGES ---
        try:
            socketio.emit('global_notification', {
                'title': '🔄 Complaint Status Changed',
                'message': f"Complaint ID ...{str(complaint.complaint_id)[-7:]} has been changed from '{old_status}' to '{new_status}' by {current_user.role} ({current_user.name})."
            }, broadcast=True)
        except Exception as e:
            print(f"WebSocket status broadcast exception: {e}")
        # ------------------------------------------------------------
        
    # Redirect cleanly based on who updated it
    if current_user.role == 'Operator':
        return redirect(url_for('operator_dashboard'))
    elif current_user.role == 'HelpDesk':
        return redirect(url_for('helpdesk_dashboard'))
    elif current_user.role == 'Admin':
        return redirect(url_for('admin_dashboard'))
    else:
        return redirect(url_for('dashboard'))
    
# ============================================
# FIELD ENGINEER - ACCEPT COMPLAINT
# ============================================

@app.route('/worker/accept/<int:comp_id>')
@login_required
def accept_complaint(comp_id):
    if current_user.role != 'FieldEngineer': 
        return redirect(url_for('dashboard'))
    
    complaint = Complaint.query.get(comp_id)
    if not complaint:
        flash('Complaint not found!', 'error')
        return redirect(url_for('worker_dashboard'))
    
    if complaint.status != 'Pending':
        flash('Complaint is not available!', 'error')
        return redirect(url_for('worker_dashboard'))
    
    # Set the accepted_at time
    complaint.assigned_to = current_user.id
    complaint.status = 'Active'
    complaint.accepted_at = datetime.now()
    complaint.assigned_at = datetime.now()
    
    db.session.commit()
    flash(f'You accepted complaint {complaint.complaint_id}!', 'success')
    
    return redirect(url_for('worker_dashboard'))

# ============================================
# FIELD ENGINEER - RESOLVE COMPLAINT
# ============================================

@app.route('/worker/resolve/<int:comp_id>', methods=['GET', 'POST'])
@login_required
def resolve_complaint(comp_id):
    if current_user.role != 'FieldEngineer': return redirect(url_for('dashboard'))
    comp = Complaint.query.get(comp_id)
    
    if request.method == 'POST':
        file = request.files.get('image')
        if file and file.filename:
            filename = f"{comp.complaint_id}_{file.filename}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            img = ComplaintImage(complaint_id=comp.id, filename=filename)
            db.session.add(img)
        
        comp.resolution_remarks = request.form['remarks']
        comp.worker_location = request.form['gps_loc']
        comp.status = 'Resolved'
        comp.resolved_at = datetime.now()
        
        # Notify HelpDesk and Operator for approval
        users = User.query.filter(User.role.in_(['HelpDesk', 'Operator'])).all()
        for user in users:
            notif = Notification(
                user_id=user.id,
                complaint_id=comp.id,
                title="Complaint Resolved - Awaiting Approval",
                message=f"{comp.complaint_id} resolved by {current_user.name}. Please approve closure."
            )
            db.session.add(notif)
        
        db.session.commit()
        flash('Complaint resolved! Waiting for approval.')
        return redirect(url_for('worker_dashboard'))
    
    return render_template('worker_resolve.html', complaint=comp)

# ============================================
# APPROVE CLOSURE (Operator & HelpDesk) - Continued
# ============================================

@app.route('/approve_close/<int:comp_id>', methods=['POST'])
@login_required
def approve_close(comp_id):
    if current_user.role not in ['Operator', 'HelpDesk']:
        return redirect(url_for('dashboard'))
    
    comp = Complaint.query.get(comp_id)
    
    # HelpDesk can only close Power Supply complaints
    if current_user.role == 'HelpDesk' and comp.category != 'Power Supply':
        flash('HelpDesk can only close Power Supply complaints!', 'error')
        return redirect(url_for('helpdesk_dashboard'))
    
    if comp and comp.status == 'Resolved':
        comp.status = 'Closed'
        comp.resolved_at = datetime.now()
        
        # Notify Field Engineer
        if comp.assigned_to:
            notif = Notification(
                user_id=comp.assigned_to,
                complaint_id=comp.id,
                title="Complaint Closed",
                message=f"Your complaint {comp.complaint_id} has been approved and closed."
            )
            db.session.add(notif)
        
        db.session.commit()
        flash(f'Complaint {comp.complaint_id} has been closed!')
    
    if current_user.role == 'Operator':
        return redirect(url_for('operator_dashboard'))
    return redirect(url_for('helpdesk_dashboard'))

# ============================================
# REJECT CLOSURE (Send back to Field Engineer)
# ============================================

@app.route('/reject_close/<int:comp_id>', methods=['POST'])
@login_required
def reject_close(comp_id):
    if current_user.role not in ['Operator', 'HelpDesk']:
        return redirect(url_for('dashboard'))
    
    comp = Complaint.query.get(comp_id)
    reason = request.form.get('reason', '')
    
    if comp and comp.status == 'Resolved':
        comp.status = 'Active'  # Send back to Field Engineer
        comp.resolved_at = None
        
        # Notify Field Engineer
        if comp.assigned_to:
            notif = Notification(
                user_id=comp.assigned_to,
                complaint_id=comp.id,
                title="Closure Rejected",
                message=f"Your complaint {comp.complaint_id} was rejected. Reason: {reason}"
            )
            db.session.add(notif)
        
        db.session.commit()
        flash(f'Complaint {comp.complaint_id} sent back to engineer!')
    
    if current_user.role == 'Operator':
        return redirect(url_for('operator_dashboard'))
    return redirect(url_for('helpdesk_dashboard'))

# ============================================
# HELPDESK - UPDATE COMPLAINT STATUS
# ============================================

@app.route('/helpdesk_update_complaint/<int:comp_id>', methods=['POST'])
@login_required
def helpdesk_update_complaint(comp_id):
    if current_user.role != 'HelpDesk':
        flash('Unauthorized access.', 'error')
        return redirect(url_for('dashboard'))

    complaint = Complaint.query.get_or_404(comp_id)
    
    new_status = request.form.get('status')
    reason_remarks = request.form.get('reason', '').strip()
    
    # Update core complaint properties
    complaint.status = new_status
    complaint.resolution_remarks = reason_remarks
    complaint.resolved_by = current_user.id  # Sets the operator tracking key
    
    if new_status in ['Active', 'Closed']:
        complaint.assigned_at = datetime.now()

    # --- AUDIT TRAIL LOGGING ---
    new_log = ComplaintLog(
        complaint_id=complaint.id,
        status=new_status,
        changed_by=current_user.id,
        reason=reason_remarks
    )
    db.session.add(new_log)
    
    try:
        db.session.commit()
        flash('Complaint status updated and logged successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash('An error occurred while updating the complaint.', 'error')
        
    return redirect(url_for('helpdesk_dashboard'))


# ============================================
# HELPDESK - CLOSE BUTTON (Simple)
# ============================================

@app.route('/helpdesk_close/<int:comp_id>', methods=['POST'])
@login_required
def helpdesk_close(comp_id):
    if current_user.role != 'HelpDesk':
        return redirect(url_for('dashboard'))
    
    complaint = Complaint.query.get_or_404(comp_id)
    
    # HelpDesk can ONLY close Power Supply complaints
    if complaint.category != 'Power Supply':
        flash(f'You can only close Power Supply complaints!', 'error')
        return redirect(url_for('helpdesk_dashboard'))
    
    # Update to Closed
    complaint.status = 'Closed'
    complaint.closed_at = datetime.now()
    complaint.resolved_by = current_user.id
    complaint.assigned_to = current_user.id
    
    db.session.commit()
    flash(f'Complaint {complaint.complaint_id} closed by {current_user.name}!', 'success')
    
    return redirect(url_for('helpdesk_dashboard'))

if __name__ == '__main__':
    # Initialize database
    with app.app_context():
        # ✅ Add missing column to existing database
        try:
            db.session.execute(text('ALTER TABLE "user" ADD COLUMN device_token VARCHAR(100);'))
            db.session.commit()  # Make sure to explicitly commit this adjustment!
            print("Column added successfully!")
        except Exception as e:
            print(f"Column already exists or error: {e}")
        
        db.create_all()
        print("Database created!")

        if not User.query.filter_by(username='admin').first():
            admin = User(
                name='System Admin',
                username='admin',
                password='ICTadmin@123',
                role='Admin',
                is_approved=True
            )
            db.session.add(admin)
            db.session.commit()

    # Create uploads folder
    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        os.makedirs(app.config['UPLOAD_FOLDER'])
    
    socketio.run(app, debug=True, port=5000)

